from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def apply_conditional_formatting():
    file_path = "data/Final Template.xlsx"

    wb = load_workbook(file_path)
    ws = wb.active

    # Define colors
    red_fill = PatternFill(patternType="solid", fgColor="FF0000")
    yellow_fill = PatternFill(patternType="solid", fgColor="FFFF00")
    green_fill = PatternFill(patternType="solid", fgColor="00FF00")

    # Loop through cells
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=9):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 15:
                    cell.fill = red_fill

                elif cell.value > 10:
                    cell.fill = yellow_fill

                else:
                    cell.fill = green_fill

    wb.save(file_path)

apply_conditional_formatting()